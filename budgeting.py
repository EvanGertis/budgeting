import openpyxl

# Create a new Excel file or open an existing one
workbook = openpyxl.Workbook()
sheet = workbook.active

# Define the questions as a list of strings
questions = [
    "Did you complete your weekly budget review and stick to your budget for the week? (Enter 1 for yes, 0 for no)",
    "Did you overspend in any categories and are there any areas where you can cut back on expenses or find ways to save money for retirement and vacation? (Enter the amount overspent or 0)",
    "Were there any unexpected expenses that came up during the week and should you consider earmarking dollars for retirement and vacation? (Enter the amount spent or 0)",
    "How much money do you have left for the rest of the month and are there any upcoming expenses that you need to plan for to ensure you’re still saving for retirement and vacation? (Enter the remaining amount or 0)",
    "Did you save any money this week, and if so, can you put it towards your retirement and vacation goals? (Enter the amount saved or 0)",
    "Where are your savings going and what are they doing to help you reach your retirement and vacation goals? (Enter the amount saved or 0)",
    "How are you managing your financial risk and protecting your retirement and vacation savings? (Enter the amount invested or 0)",
    "Do you have adequate retirement and vacation savings and different accounts designed for different periods of life? (Enter 1 for yes, 0 for no)",
    "How does your spending this week compare to previous weeks or months in terms of saving for retirement and vacation? (Enter the percentage change or 0)",
    "Are you being strategic about the future by allocating dollars correctly to get more for retirement and vacation? (Enter 1 for yes, 0 for no)"
]

# Define a function to ask the user each question and store their answer in a variable
def ask_question(question):
    while True:
        try:
            answer = float(input(question))
            return answer
        except ValueError:
            print("Please enter a numeric value.")

# Call the function for each question and store the answers in a list
answers = []
for question in questions:
    answer = ask_question(question)
    answers.append(answer)

# Populate the Excel sheet with the results by writing the list of questions and answers to the sheet
for i in range(len(questions)):
    sheet.cell(row=i+1, column=1, value=questions[i])
    sheet.cell(row=i+1, column=2, value=answers[i])

# Save the Excel file
workbook.save("results.xlsx")